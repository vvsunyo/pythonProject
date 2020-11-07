import openpyxl
import excelUtil

if __name__ == '__main__':
    wb = openpyxl.Workbook()
    wb.create_sheet("testyao")
    wb.save('cases.xlsx')
    wb = openpyxl.load_workbook("cases.xlsx")
    sheets = wb.get_sheet_by_name()

    for i in range(len(sheets)):
        if sheets[i].find("发货明细") != -1:
            sheet = wb.get_sheet_by_name(sheets[i])
            rowValue = excelUtil.excel.getRowValues(sheet,1)

        else:
            print("no target in sheet")